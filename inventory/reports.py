from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import os
from .models import Chemicals, ChemicalActivity, Locations
from drf_spectacular.utils import extend_schema, OpenApiParameter
from drf_spectacular.types import OpenApiTypes

# Constants for styling
PDF_TITLE_STYLE = ParagraphStyle(
    name='Title',
    fontName='Helvetica-Bold',
    fontSize=16,
    alignment=1,
    spaceAfter=12
)

# Helper function to create report headers
def get_report_header(title, start_date, end_date):
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    elements.append(Paragraph(title, PDF_TITLE_STYLE))
    
    # Date range
    date_text = f"Period: {start_date} to {end_date}"
    elements.append(Paragraph(date_text, styles['Normal']))
    elements.append(Spacer(1, 0.25 * inch))
    
    return elements

# Common PDF generation function with improved styling
def generate_pdf_report(title, headers, rows, start_date, end_date):
    buffer = BytesIO()
    
    # Use landscape for wider tables
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                           leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    
    # Start with header elements
    elements = get_report_header(title, start_date, end_date)
    
    # Create the table with data
    if rows:
        table = Table([headers] + rows, repeatRows=1)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
    else:
        # Add a message if no data
        styles = getSampleStyleSheet()
        elements.append(Paragraph("No data available for the selected period.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.pdf"'
    response.write(pdf)
    
    return response

# Excel generation with improved styling and formatting
def generate_excel_report(title, headers, rows, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet names limited to 31 chars
    
    # Add title and date range
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    date_range = f"Period: {start_date} to {end_date}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=2, value=date_range)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Style for headers
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    
    # Write headers
    header_row = 4  # Start at row 4 to leave space for title
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Data styling
    data_alignment = Alignment(vertical='center', wrap_text=True)
    data_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    
    # Write data with alternating row colors
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    for row_idx, row in enumerate(rows, header_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = data_border
            
            # Apply alternating row colors
            if row_idx % 2 == 1:
                cell.fill = alt_fill
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_cells = [ws.cell(row=i, column=col).value for i in range(4, len(rows) + header_row + 1)]
        column_cells = [str(cell) if cell is not None else '' for cell in column_cells]
        max_length = max([len(headers[col-1])] + [len(str(cell)) for cell in column_cells])
        adjusted_width = max_length + 4  # Add padding
        ws.column_dimensions[get_column_letter(col)].width = min(adjusted_width, 40)  # Cap width at 40
    
    # Freeze header row
    ws.freeze_panes = 'A5'
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    
    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.xlsx"'
    
    return response

@extend_schema(
    tags=['Reports'],
    description='Generate inventory report',
    parameters=[
        OpenApiParameter('format', OpenApiTypes.STR, enum=['pdf', 'excel'], 
                        description='Report format (pdf or excel)'),
        OpenApiParameter('start_date', OpenApiTypes.DATE, 
                        description='Start date for report range'),
        OpenApiParameter('end_date', OpenApiTypes.DATE, 
                        description='End date for report range'),
        OpenApiParameter('location', OpenApiTypes.STR, 
                        description='Filter by location ID (optional)'),
    ],
    responses={
        200: {'type': 'string', 'format': 'binary'},
        400: {'description': 'Invalid parameters'},
        500: {'description': 'Server error'}
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def inventory_report(request):
    """
    Generate a comprehensive inventory report with current stock levels.
    """
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    export_format = request.query_params.get('format', 'pdf')
    location_id = request.query_params.get('location')
    
    # Validate parameters
    if not all([start_date, end_date]):
        return Response({"error": "Missing required parameters: start_date and end_date"}, status=400)
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)
    
    try:
        # Build query for chemicals
        query = Chemicals.objects.all().select_related('location', 'created_by')
        
        # Apply location filter if provided
        if location_id:
            query = query.filter(location_id=location_id)
        
        # Prepare the report data
        headers = ['Chemical Name', 'Formula', 'Location', 'Quantity', 'State', 'Type', 'Hazard Info', 'Expiry Date', 'Last Updated']
        rows = []
        
        for chemical in query:
            unit_suffix = "L" if chemical.chemical_state == "Liquid" else "g"
            rows.append([
                chemical.name,
                chemical.molecular_formula,
                chemical.location.name,
                f"{chemical.quantity} {unit_suffix}",
                chemical.chemical_state,
                chemical.chemical_type,
                chemical.hazard_information[:50] + '...' if len(chemical.hazard_information) > 50 else chemical.hazard_information,
                chemical.expires.strftime('%Y-%m-%d') if chemical.expires else 'N/A',
                chemical.updated_at.strftime('%Y-%m-%d') if chemical.updated_at else 'N/A'
            ])
        
        # Generate report based on format
        title = 'Chemical Inventory Report'
        if export_format.lower() == 'pdf':
            return generate_pdf_report(title, headers, rows, start_date, end_date)
        else:
            return generate_excel_report(title, headers, rows, start_date, end_date)
    except Exception as e:
        import traceback
        print(f"Error generating inventory report: {str(e)}")
        print(traceback.format_exc())
        return Response({"error": f"Error generating report: {str(e)}"}, status=500)

@extend_schema(
    tags=['Reports'],
    description='Generate usage report',
    parameters=[
        OpenApiParameter('format', OpenApiTypes.STR, enum=['pdf', 'excel'], 
                        description='Report format (pdf or excel)'),
        OpenApiParameter('start_date', OpenApiTypes.DATE, 
                        description='Start date for report range'),
        OpenApiParameter('end_date', OpenApiTypes.DATE, 
                        description='End date for report range'),
        OpenApiParameter('chemical_id', OpenApiTypes.STR, 
                        description='Filter by chemical ID (optional)'),
        OpenApiParameter('user_id', OpenApiTypes.STR, 
                        description='Filter by user ID (optional)'),
    ],
    responses={
        200: {'type': 'string', 'format': 'binary'},
        400: {'description': 'Invalid parameters'},
        500: {'description': 'Server error'}
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def usage_report(request):
    """
    Generate a detailed usage report showing all chemical activities.
    """
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    export_format = request.query_params.get('format', 'pdf')
    chemical_id = request.query_params.get('chemical_id')
    user_id = request.query_params.get('user_id')
    
    # Validate parameters
    if not all([start_date, end_date]):
        return Response({"error": "Missing required parameters: start_date and end_date"}, status=400)
    
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        # Adjust end_date to include the entire day
        end_date_obj = datetime.combine(end_date_obj, datetime.max.time())
    except ValueError:
        return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)
    
    try:
        # Build query for chemical activities
        query = ChemicalActivity.objects.filter(
            timestamp__range=[start_date_obj, end_date_obj]
        ).select_related('chemical', 'user', 'chemical__location')
        
        # Apply filters if provided
        if chemical_id:
            query = query.filter(chemical_id=chemical_id)
        
        if user_id:
            query = query.filter(user_id=user_id)
        
        # Prepare the report data
        headers = ['Date & Time', 'Chemical', 'Action', 'Quantity', 'Location', 'User', 'Notes']
        rows = []
        
        for activity in query:
            unit_suffix = "L" if activity.chemical.chemical_state == "Liquid" else "g"
            rows.append([
                activity.timestamp.strftime('%Y-%m-%d %H:%M'),
                activity.chemical.name,
                activity.action.title(),
                f"{abs(activity.quantity)} {unit_suffix}",
                activity.chemical.location.name,
                activity.user.get_full_name(),
                activity.notes if activity.notes else 'N/A'
            ])
        
        # Sort by date (newest first)
        rows.sort(key=lambda x: x[0], reverse=True)
        
        # Generate report based on format
        title = 'Chemical Usage Report'
        if export_format.lower() == 'pdf':
            return generate_pdf_report(title, headers, rows, start_date, end_date)
        else:
            return generate_excel_report(title, headers, rows, start_date, end_date)
    except Exception as e:
        import traceback
        print(f"Error generating usage report: {str(e)}")
        print(traceback.format_exc())
        return Response({"error": f"Error generating report: {str(e)}"}, status=500)

@extend_schema(
    tags=['Reports'],
    description='Generate expiry report',
    parameters=[
        OpenApiParameter('format', OpenApiTypes.STR, enum=['pdf', 'excel'], 
                        description='Report format (pdf or excel)'),
        OpenApiParameter('days', OpenApiTypes.INT, 
                        description='Number of days to look ahead (default: 90)'),
    ],
    responses={
        200: {'type': 'string', 'format': 'binary'},
        400: {'description': 'Invalid parameters'},
        500: {'description': 'Server error'}
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def expiry_report(request):
    """
    Generate a report of chemicals that will expire soon.
    """
    export_format = request.query_params.get('format', 'pdf')
    days_ahead = request.query_params.get('days', '90')
    
    try:
        days_ahead = int(days_ahead)
        if days_ahead <= 0:
            return Response({"error": "Days parameter must be a positive number"}, status=400)
    except ValueError:
        return Response({"error": "Days parameter must be a valid number"}, status=400)
    
    try:
        # Calculate date range
        today = timezone.now().date()
        expiry_cutoff = today + timedelta(days=days_ahead)
        
        # Build query for chemicals expiring soon
        query = Chemicals.objects.filter(
            expires__range=[today, expiry_cutoff]
        ).select_related('location', 'created_by')
        
        # Prepare the report data
        headers = ['Chemical Name', 'Location', 'Quantity', 'Expiry Date', 'Days Left', 'Added By', 'Creation Date']
        rows = []
        
        for chemical in query:
            unit_suffix = "L" if chemical.chemical_state == "Liquid" else "g"
            days_left = (chemical.expires - today).days
            
            rows.append([
                chemical.name,
                chemical.location.name,
                f"{chemical.quantity} {unit_suffix}",
                chemical.expires.strftime('%Y-%m-%d'),
                str(days_left),
                chemical.created_by.get_full_name(),
                chemical.created_at.strftime('%Y-%m-%d')
            ])
        
        # Sort by days_left (ascending)
        rows.sort(key=lambda x: int(x[4]))
        
        # Generate report based on format
        title = f'Chemicals Expiring Within {days_ahead} Days'
        if export_format.lower() == 'pdf':
            return generate_pdf_report(title, headers, rows, today.strftime('%Y-%m-%d'), expiry_cutoff.strftime('%Y-%m-%d'))
        else:
            return generate_excel_report(title, headers, rows, today.strftime('%Y-%m-%d'), expiry_cutoff.strftime('%Y-%m-%d'))
    except Exception as e:
        import traceback
        print(f"Error generating expiry report: {str(e)}")
        print(traceback.format_exc())
        return Response({"error": f"Error generating report: {str(e)}"}, status=500)

@extend_schema(
    tags=['Reports'],
    description='Generate stock levels report',
    parameters=[
        OpenApiParameter('format', OpenApiTypes.STR, enum=['pdf', 'excel'], 
                        description='Report format (pdf or excel)'),
        OpenApiParameter('threshold', OpenApiTypes.FLOAT, 
                        description='Quantity threshold for low stock (default: 100)'),
    ],
    responses={
        200: {'type': 'string', 'format': 'binary'},
        400: {'description': 'Invalid parameters'},
        500: {'description': 'Server error'}
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def low_stock_report(request):
    """
    Generate a report of chemicals with low stock levels.
    """
    export_format = request.query_params.get('format', 'pdf')
    threshold = request.query_params.get('threshold', '100')
    
    try:
        threshold = float(threshold)
        if threshold <= 0:
            return Response({"error": "Threshold parameter must be a positive number"}, status=400)
    except ValueError:
        return Response({"error": "Threshold parameter must be a valid number"}, status=400)
    
    try:
        # Build query for chemicals with low stock
        query = Chemicals.objects.filter(
            quantity__lte=threshold
        ).select_related('location')
        
        # Prepare the report data
        headers = ['Chemical Name', 'Formula', 'Location', 'Current Stock', 'State', 'Expiry Date']
        rows = []
        
        today = timezone.now().date()
        
        for chemical in query:
            unit_suffix = "L" if chemical.chemical_state == "Liquid" else "g"
            
            rows.append([
                chemical.name,
                chemical.molecular_formula,
                chemical.location.name,
                f"{chemical.quantity} {unit_suffix}",
                chemical.chemical_state,
                chemical.expires.strftime('%Y-%m-%d') if chemical.expires else 'N/A'
            ])
        
        # Sort by current stock (ascending)
        rows.sort(key=lambda x: float(x[3].split()[0]))
        
        # Generate report based on format
        title = f'Chemicals Below Stock Threshold ({threshold})'
        if export_format.lower() == 'pdf':
            return generate_pdf_report(title, headers, rows, today.strftime('%Y-%m-%d'), today.strftime('%Y-%m-%d'))
        else:
            return generate_excel_report(title, headers, rows, today.strftime('%Y-%m-%d'), today.strftime('%Y-%m-%d'))
    except Exception as e:
        import traceback
        print(f"Error generating low stock report: {str(e)}")
        print(traceback.format_exc())
        return Response({"error": f"Error generating report: {str(e)}"}, status=500) 